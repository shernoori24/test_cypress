# app/services/data_service.py
# Ce fichier contient le service principal de gestion des données
# Il gère les opérations CRUD sur les fichiers Excel (inscriptions/présences)
# et les fonctionnalités de planification/reporting

import pandas as pd              # Bibliothèque pour manipuler les données (DataFrames)
from pathlib import Path         # Pour gérer les chemins de fichiers de manière moderne
import openpyxl                  # Pour manipuler les fichiers Excel directement
import os                        # Pour les opérations système
from datetime import datetime, timedelta    # Pour gérer les dates et heures
import json                      # Pour manipuler les fichiers JSON
import re                        # Pour les expressions régulières
import logging                   # Pour les logs
from .data_cleaning_service import DataCleaningService  # Service de nettoyage des données
from .file_lock_service import file_lock_service  # Service de verrouillage de fichiers

# Configuration du logger pour ce module
logger = logging.getLogger(__name__)

class DataService:
    """
    Service principal pour la gestion des données de l'application
    
    Cette classe centralise toutes les opérations sur les données :
    - Chargement des fichiers Excel (inscriptions et présences)
    - Ajout/modification des inscriptions et présences
    - Génération de rapports et statistiques
    - Gestion du planning hebdomadaire
    - Export/import de données
    
    Pattern utilisé : Service Layer + Data Access Object (DAO)
    - Service Layer : Logique métier centralisée
    - DAO : Accès aux données (fichiers Excel, JSON)
    """
    
    def __init__(self, data_folder='data'):
        """
        Initialise le service de données
        
        Args:
            data_folder (str): Nom du dossier contenant les fichiers de données
        """
        # Utiliser la même logique de chemin que Config.py pour la cohérence en mode production
        import sys
        if getattr(sys, 'frozen', False):
            # Mode production (PyInstaller) - utiliser le dossier de l'exécutable
            base_dir = Path(sys.executable).parent
        else:
            # Mode développement - utiliser le dossier du projet (3 niveaux au-dessus)
            base_dir = Path(__file__).parent.parent.parent
        
        self.data_folder = base_dir / data_folder
        
        # Initialiser le data_loader pour utiliser le mécanisme de vérification de modification de fichiers
        from app import data_loader
        self.data_loader = data_loader
        
        # Initialiser le service de nettoyage
        self.cleaning_service = DataCleaningService()
        
        # Variables pour le suivi de la qualité des données
        self.last_cleaning_alerts = {'inscription': [], 'presence': []}
        self.data_quality_score = 0
        
        # Charger les données au démarrage
        self.load_data()
    
    def load_data(self):
        """
        Charge les données depuis les fichiers Excel avec nettoyage automatique
        
        Cette méthode utilise un try/except pour gérer le cas où
        les fichiers n'existent pas encore (première utilisation)
        Applique automatiquement le nettoyage des données pour garantir la qualité
        """
        try:
            # Utiliser le DataLoader pour obtenir les DataFrames
            # Le DataLoader gère automatiquement le cache et la vérification de modification de fichiers
            raw_inscription_df = self.data_loader.inscriptions.copy()
            raw_presence_df = self.data_loader.presences.copy()
            
            # Appliquer le nettoyage automatique
            logger.debug("Nettoyage automatique des données en cours...")
            
            # Nettoyer les données d'inscription
            self.inscription_df, inscription_alerts = self.cleaning_service.clean_inscription_data(raw_inscription_df)
            self.last_cleaning_alerts['inscription'] = inscription_alerts
            
            # Nettoyer les données de présence
            self.presence_df, presence_alerts = self.cleaning_service.clean_presence_data(raw_presence_df)
            self.last_cleaning_alerts['presence'] = presence_alerts
            
            # Calculer le score de qualité des données
            total_alerts = len(inscription_alerts) + len(presence_alerts)
            total_records = len(self.inscription_df) + len(self.presence_df)
            self.data_quality_score = max(0, 100 - (total_alerts / max(total_records, 1) * 100))
            
            # Log du résumé de nettoyage (seulement en mode debug)
            logger.debug(f"Nettoyage terminé - Score qualité: {self.data_quality_score:.1f}%")
            if total_alerts > 0:
                logger.debug(f"   {total_alerts} problèmes détectés et corrigés")
                logger.debug(f"   Inscriptions: {len(inscription_alerts)} alertes")
                logger.debug(f"   Présences: {len(presence_alerts)} alertes")
            
        except Exception as e:
            print(f"❌ Erreur lors du chargement des données: {e}")
            # Créer des DataFrames vides pour éviter les erreurs
            self.inscription_df = pd.DataFrame()
            self.presence_df = pd.DataFrame()
            self.last_cleaning_alerts = {'inscription': [f'Erreur: {str(e)}'], 'presence': []}
            self.data_quality_score = 0
    
    def get_total_inscrits(self):
        """
        Retourne le nombre total d'apprenants inscrits
        
        Returns:
            int: Nombre total de lignes dans le DataFrame des inscriptions
        """
        return self.inscription_df.shape[0]

    

    def get_presence_days_for_person(self, person_name):
        """
        Retourne les jours de présence d'une personne spécifique
        
        Cette méthode recherche une personne par son nom dans les données
        de présence et retourne les jours de la semaine où elle était présente
        
        Args:
            person_name (str): Nom de la personne à rechercher
            
        Returns:
            list or None: Liste des jours de la semaine (en anglais) ou None si pas trouvé
        """
        # Vérifier que les DataFrames ne sont pas vides
        if self.presence_df.empty or self.inscription_df.empty:
            return None

        # Copier les DataFrames pour éviter de modifier les originaux
        presence_df = self.presence_df.copy()
        inscription_df = self.inscription_df.copy()

        # Créer une colonne nom complet dans les inscriptions
        # pour faciliter la correspondance avec les présences
        inscription_df['FullName'] = inscription_df['Prénom'].fillna('') + ' ' + inscription_df['NOM'].fillna('')

        # Vérifier que le nom de la personne est fourni
        if not person_name:
            return None

        # Échapper les caractères spéciaux pour éviter les erreurs regex
        safe_person_name = re.escape(person_name)

        # Filtrer les présences où le nom de l'apprenant contient le nom recherché
        filtered_presence = presence_df[presence_df['Apprenant'].str.contains(safe_person_name, case=False, na=False)]

        if filtered_presence.empty:
            return None

        # Fusionner les présences filtrées avec les inscriptions
        merged_df = filtered_presence.merge(inscription_df, left_on='Apprenant', right_on='FullName', how='left')

        if merged_df.empty:
            return None

        # Extraire le jour de la semaine depuis la colonne 'Date du Jour'
        merged_df.loc[:, 'Jour de la semaine'] = pd.to_datetime(merged_df['Date du Jour']).dt.day_name()

        # Obtenir les jours uniques de présence
        days_present = merged_df['Jour de la semaine'].unique().tolist()

        return days_present

    def add_inscription(self, data):
        try:
            file_path = self.data_folder / 'inscription.xlsx'
            
            # Calculer l'âge si la date de naissance est fournie
            age = ""
            if data.get('date_naissance'):
                birth_date = datetime.strptime(data['date_naissance'], '%Y-%m-%d')
                today = datetime.now()
                age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                age = f"{age} ans"
            
            # Générer un numéro d'inscription automatique
            current_year = datetime.now().year % 100  # Derniers 2 chiffres de l'année
            numero = f"{current_year:02d}-"
            
            if file_path.exists():
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Trouver le prochain numéro disponible
                existing_numbers = []
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                    if row[0] and isinstance(row[0], str) and row[0].startswith(f"{current_year:02d}-"):
                        try:
                            num = int(row[0].split('-')[1])
                            existing_numbers.append(num)
                        except (ValueError, IndexError):
                            continue
                
                next_number = max(existing_numbers, default=0) + 1
                numero += f"{next_number:03d}"
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                # Créer les en-têtes selon la structure Excel
                headers = [
                    "N°", "NOM", "Prénom", "Sexe", "Date de naissance", "Age", 
                    "Pays de naissance", "Nationalité", "Continent", "ISO", 
                    "Arrivée en France", "Document", "Statut à l'entrée", "Statut actuel",
                    "Adresse", "Code postal", "Ville", "Prioritaire/Veille", 
                    "Type de logement", "Téléphone", "Email", "Situation Familiale",
                    "Revenus", "Langues Parlées", "Prescripteur", "Structure actuelle",
                    "Date inscription", "Première venue", "Commentaires"
                ]
                ws.append(headers)
                numero += "001"
            
            # Préparer les données dans l'ordre des colonnes Excel
            row_data = [
                numero,                              # N°
                data.get('nom', ''),                # NOM
                data.get('prenom', ''),             # Prénom
                data.get('sexe', ''),               # Sexe
                data.get('date_naissance', ''),     # Date de naissance
                age,                                # Age
                data.get('pays_naissance', ''),     # Pays de naissance
                data.get('nationalite', ''),        # Nationalité
                data.get('continent', ''),          # Continent
                '',                                 # ISO (vide par défaut)
                data.get('arrivee_france', ''),     # Arrivée en France
                data.get('document', ''),           # Document
                data.get('statut_entree', ''),      # Statut à l'entrée
                data.get('statut_actuel', ''),      # Statut actuel
                data.get('adresse', ''),            # Adresse
                data.get('code_postal', ''),        # Code postal
                data.get('ville', ''),              # Ville
                data.get('prioritaire_veille', ''), # Prioritaire/Veille
                data.get('type_logement', ''),      # Type de logement
                data.get('telephone', ''),          # Téléphone
                data.get('email', ''),              # Email
                data.get('situation_familiale', ''), # Situation Familiale
                data.get('revenus', ''),            # Revenus
                data.get('langues_parlees', ''),    # Langues Parlées
                data.get('prescripteur', ''),       # Prescripteur
                data.get('structure_actuelle', ''), # Structure actuelle
                data.get('date_inscription', ''),   # Date inscription
                data.get('premiere_venue', ''),     # Première venue
                data.get('commentaires', '')        # Commentaires
            ]
            
            ws.append(row_data)
            wb.save(file_path)
            self.load_data()  # Reload data after update
            return True, f"Inscription ajoutée avec succès ! Numéro d'inscription: {numero}"
        except Exception as e:
            return False, f"Erreur lors de l'ajout : {str(e)}"


    def get_person_info_by_name(self, person_name):
        """Get formatted information about a person by their name."""
        if self.inscription_df.empty:
            return None

        inscription_df = self.inscription_df.copy()

        # Filter rows where NOM or Prénom contains person_name (case insensitive)
        filtered = inscription_df[
            (inscription_df['NOM'].str.contains(person_name, case=False, na=False)) |
            (inscription_df['Prénom'].str.contains(person_name, case=False, na=False))
        ]

        if filtered.empty:
            return None

        # Take the first matching record and format info
        person = filtered.iloc[0]

        info = (
            f"NOM: {person.get('NOM', '')}, "
            f"Prénom: {person.get('Prénom', '')}, "
            f"Adresse: {person.get('Adresse', '')}, "
            f"Code postal: {person.get('Code postal', '')}, "
            f"Ville: {person.get('Ville', '')}, "
            f"Téléphone: {person.get('Téléphone', '')}, "
            f"Email: {person.get('Email', '')}"
        )
        return info

    def get_filtered_inscriptions(self, search='', sexe='', ville='', prioritaire=''):
        """Retourner les inscriptions filtrées selon les critères"""
        if self.inscription_df.empty:
            return []
        
        df = self.inscription_df.copy()
        
        # Filtre par recherche (nom ou prénom)
        if search:
            mask = (df['NOM'].str.contains(search, case=False, na=False) | 
                   df['Prénom'].str.contains(search, case=False, na=False))
            df = df[mask]
        
        # Filtre par sexe
        if sexe:
            df = df[df['Sexe'] == sexe]
        
        # Filtre par ville
        if ville:
            df = df[df['Ville'] == ville]
        
        # Filtre par prioritaire
        if prioritaire:
            df = df[df['Prioritaire/Veille'] == prioritaire]
        
        return df.to_dict('records')
    
    def get_inscription_statistics(self):
        """Calculer les statistiques des inscriptions"""
        if self.inscription_df.empty:
            return {'total': 0, 'mois': 0, 'semaine': 0, 'prioritaires': 0}
        
        df = self.inscription_df.copy()
        
        # Total
        total = len(df)
        
        # Ce mois-ci
        current_month = datetime.now().strftime('%m/%Y')
        mois = 0
        if 'Date inscription' in df.columns:
            try:
                # Convertir les dates et filtrer
                df['Date inscription'] = pd.to_datetime(df['Date inscription'], errors='coerce')
                current_month_mask = df['Date inscription'].dt.strftime('%m/%Y') == current_month
                mois = current_month_mask.sum()
            except:
                mois = 0
        
        # Cette semaine (approximation)
        semaine = 0
        if 'Date inscription' in df.columns:
            try:
                week_start = datetime.now() - pd.Timedelta(days=7)
                recent_mask = df['Date inscription'] >= week_start
                semaine = recent_mask.sum()
            except:
                semaine = 0
        
        # Prioritaires
        prioritaires = 0
        if 'Prioritaire/Veille' in df.columns:
            prioritaires = (df['Prioritaire/Veille'] == 'P').sum()
        
        return {
            'total': total,
            'mois': mois,
            'semaine': semaine,
            'prioritaires': prioritaires
        }
    
    def get_unique_villes(self):
        """Retourner la liste unique des villes"""
        if self.inscription_df.empty or 'Ville' not in self.inscription_df.columns:
            return []
        
        return sorted(self.inscription_df['Ville'].dropna().unique().tolist())
    
    def get_inscription_details(self, numero):
        """Retourner les détails d'une inscription par son numéro"""
        if self.inscription_df.empty:
            return None
        
        inscription = self.inscription_df[self.inscription_df['N°'] == numero]
        if inscription.empty:
            return None
        
        return inscription.iloc[0].to_dict()
    
    def export_inscriptions(self):
        """Créer un fichier Excel d'export"""
        import tempfile
        import shutil
        
        # Créer une copie du fichier d'origine
        original_path = self.data_folder / 'inscription.xlsx'
        
        # Créer un fichier temporaire
        temp_dir = tempfile.mkdtemp()
        export_path = Path(temp_dir) / 'inscriptions_export.xlsx'
        
        if original_path.exists():
            shutil.copy2(original_path, export_path)
        else:
            # Créer un nouveau fichier vide si l'original n'existe pas
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Aucune donnée disponible"])
            wb.save(export_path)
        
        return export_path

    def get_all_apprenants(self):
        """Retourner la liste de tous les apprenants pour le sélecteur, triés par date d'inscription (plus récents en premier)"""
        if self.inscription_df.empty:
            return []
        
        apprenants = []
        for _, row in self.inscription_df.iterrows():
            apprenant_data = {
                'numero': row['N°'],
                'nom': row.get('NOM', ''),
                'prenom': row.get('Prénom', ''),
                'nom_complet': f"{row.get('NOM', '')} {row.get('Prénom', '')}"  # MODIFIÉ: NOM Prénom
            }
            
            # Ajouter la date d'inscription si disponible pour le tri
            if 'Date inscription' in self.inscription_df.columns:
                try:
                    date_inscription = pd.to_datetime(row.get('Date inscription'), errors='coerce')
                    apprenant_data['date_inscription'] = date_inscription
                except:
                    apprenant_data['date_inscription'] = pd.NaT
            else:
                apprenant_data['date_inscription'] = pd.NaT
            
            apprenants.append(apprenant_data)
        
        # Trier par date d'inscription (plus récents en premier), puis par nom complet
        # Les apprenants sans date valide seront placés à la fin
        apprenants.sort(key=lambda x: (
            x['date_inscription'] is pd.NaT,  # False pour les dates valides (placées en premier)
            -x['date_inscription'].timestamp() if x['date_inscription'] is not pd.NaT else 0,  # Tri décroissant par timestamp
            x['nom_complet']  # Tri alphabétique en cas d'égalité
        ))
        
        # Retirer la date du résultat final pour garder la même structure
        for apprenant in apprenants:
            del apprenant['date_inscription']
        
        return apprenants
    
    def get_apprenants_list(self):
        """Alias pour get_all_apprenants - compatibilité"""
        return self.get_all_apprenants()
    
    def get_apprenant_complete_info(self, numero_apprenant, dates_filter=None):
        """Retourner toutes les informations d'un apprenant avec validation et nettoyage, avec filtrage optionnel par dates"""
        if self.inscription_df.empty:
            return None
        
        # Valider l'existence de l'apprenant
        validation = self.cleaning_service.validate_apprenant_exists(
            numero_apprenant, self.inscription_df, self.presence_df
        )
        
        # Informations d'inscription
        inscription = self.inscription_df[self.inscription_df['N°'] == numero_apprenant]
        if inscription.empty:
            return {
                'error': 'Apprenant non trouvé dans les inscriptions',
                'validation': validation
            }
        
        # Nettoyer les valeurs NaN pour éviter les erreurs JSON
        apprenant_info = inscription.iloc[0].to_dict()
        apprenant_info_clean = {}
        for key, value in apprenant_info.items():
            if pd.isna(value):
                apprenant_info_clean[key] = ''
            elif isinstance(value, (int, float)) and pd.isna(value):
                apprenant_info_clean[key] = 0
            else:
                apprenant_info_clean[key] = str(value) if not isinstance(value, (int, float, bool)) else value
        
        # Informations de présence avec filtrage optionnel
        presences = []
        presence_status = "no_data"
        
        if not self.presence_df.empty and validation['exists_presence']:
            presence_apprenant = self.presence_df[self.presence_df['Numéro Apprenant'] == numero_apprenant]
            
            # Appliquer le filtrage par dates si demandé
            if dates_filter:
                presence_apprenant = self._filter_presences_by_date(presence_apprenant, dates_filter)
            
            presences_raw = presence_apprenant.to_dict('records')
            
            # Nettoyer les présences aussi
            for presence in presences_raw:
                presence_clean = {}
                for key, value in presence.items():
                    if pd.isna(value):
                        presence_clean[key] = ''
                    elif isinstance(value, (int, float)) and pd.isna(value):
                        presence_clean[key] = 0
                    else:
                        presence_clean[key] = str(value) if not isinstance(value, (int, float, bool)) else value
                presences.append(presence_clean)
            
            presence_status = "has_data" if presences else "no_presence"
        else:
            presence_status = "no_presence"
        
        # Statistiques de présence
        total_presences = len(presences)
        dates_uniques = []
        activites_uniques = set()
        total_heures_minutes = 0  # Total en minutes
        
        if presences:
            # Extraire les dates uniques et calculer les heures totales
            for p in presences:
                if 'Date du Jour' in p and p['Date du Jour'] and p['Date du Jour'] != '':
                    dates_uniques.append(p['Date du Jour'])
                if 'Activités' in p and p['Activités'] and p['Activités'] != '':
                    activites_uniques.add(p['Activités'])
                
                # Calculer la durée en minutes
                if 'Durée Activité Apprenants' in p and p['Durée Activité Apprenants']:
                    duree_str = str(p['Durée Activité Apprenants'])
                    total_heures_minutes += self._convertir_duree_en_minutes(duree_str)
        
        nb_jours_uniques = len(set(dates_uniques))
        
        # Calculer le taux de présence (basé sur un mois de 22 jours ouvrables)
        jours_ouvrables_mois = 22
        taux_presence = (nb_jours_uniques / jours_ouvrables_mois * 100) if nb_jours_uniques > 0 else 0
        taux_presence = min(100, taux_presence)  # Plafonner à 100%
        
        # Convertir les minutes en format heures
        total_heures_format = self._convertir_minutes_en_heures(total_heures_minutes)
        
        result = {
            'apprenant': apprenant_info_clean,
            'presences': presences,
            'presence_status': presence_status,
            'validation': validation,
            'statistiques': {
                'total_presences': total_presences,
                'jours_uniques': nb_jours_uniques,
                'activites_uniques': list(activites_uniques),
                'taux_presence': round(taux_presence, 1),
                'total_heures': total_heures_format,
                'total_minutes': total_heures_minutes
            },
            'data_quality': {
                'score': self.data_quality_score,
                'alerts': self.last_cleaning_alerts,
                'cleaned_at': datetime.now().isoformat()
            }
        }

        # === Lire les évaluations depuis grille_suivit_apprenant.xlsx ===
        try:
            eval_path = self.data_folder / 'grille_suivit_apprenant.xlsx'
            evaluations_list = []
            if eval_path.exists():
                eval_df = pd.read_excel(eval_path)

                # Normaliser les noms de colonnes pour faciliter la correspondance
                def normalize_col(c):
                    if not isinstance(c, str):
                        return str(c)
                    return c.strip().replace('\n', '').replace(' ', '_')

                eval_df.columns = [normalize_col(c) for c in eval_df.columns]

                # Colonnes attendues (apporter des alternatives si noms différents)
                col_candidates = {
                    'numero': ['Numero_apprenant', 'Numero', 'Numero_apprenant'.replace('\n','')],
                    'date': ['Date', 'date', 'DATE'],
                    'note': ['Note_test_sur_20', 'Note', 'Note_sur_20'],
                    'compr_oral': ['Comprehension_orale', 'Comprehension_orale'.replace('\n','')],
                    'compr_ecrit': ['Comprehension_ecrit', 'Comprehension_ecrit'.replace('\n',''), 'Comprehension_écrit', 'Comprehension_ecrit'],
                    'prod_oral': ['Production_orale', 'Production_orale'.replace('\n','')],
                    'prod_ecrit': ['Production_ecrit', 'Production_ecrit'.replace('\n','')],
                    'observation': ['Observation', 'Observations'],
                    'niveau': ['Niveau_global', 'Niveau']
                }

                # Helper to find first existing column name
                def find_col(df_cols, candidates):
                    for c in candidates:
                        if c in df_cols:
                            return c
                    return None

                df_cols = set(eval_df.columns)
                # Build a mapping from canonical keys to actual df columns
                mapping = {}
                mapping['numero'] = find_col(df_cols, col_candidates['numero'])
                mapping['date'] = find_col(df_cols, col_candidates['date'])
                mapping['note'] = find_col(df_cols, col_candidates['note'])
                mapping['compr_oral'] = find_col(df_cols, col_candidates['compr_oral'])
                mapping['compr_ecrit'] = find_col(df_cols, col_candidates['compr_ecrit'])
                mapping['prod_oral'] = find_col(df_cols, col_candidates['prod_oral'])
                mapping['prod_ecrit'] = find_col(df_cols, col_candidates['prod_ecrit'])
                mapping['observation'] = find_col(df_cols, col_candidates['observation'])
                mapping['niveau'] = find_col(df_cols, col_candidates['niveau'])

                # Filtrer par numéro d'apprenant (comparer en str pour robustesse)
                if mapping['numero']:
                    # Convertir la colonne en chaîne pour comparaison
                    eval_df[mapping['numero']] = eval_df[mapping['numero']].astype(str)
                    mask = eval_df[mapping['numero']] == str(numero_apprenant)
                    eval_filtered = eval_df[mask]
                else:
                    eval_filtered = pd.DataFrame()

                # Appliquer filtrage par dates si demandé
                if not eval_filtered.empty and dates_filter and mapping['date']:
                    try:
                        eval_filtered['__DateConverted'] = pd.to_datetime(eval_filtered[mapping['date']], errors='coerce')
                        if 'debut' in dates_filter:
                            debut = pd.to_datetime(dates_filter['debut'])
                            eval_filtered = eval_filtered[eval_filtered['__DateConverted'] >= debut]
                        if 'fin' in dates_filter:
                            fin = pd.to_datetime(dates_filter['fin'])
                            eval_filtered = eval_filtered[eval_filtered['__DateConverted'] <= fin]
                        eval_filtered = eval_filtered.drop(columns=['__DateConverted'])
                    except Exception:
                        # En cas d'erreur de parsing, ignorer le filtrage
                        pass

                # Construire la liste d'évaluations avec les clés attendues côté front
                for _, row in eval_filtered.iterrows():
                    ev = {
                        'Date': str(row[mapping['date']]) if mapping.get('date') and pd.notna(row.get(mapping['date'])) else '',
                        'Note_test_sur_20': row.get(mapping['note'], '') if mapping.get('note') else '',
                        'Comprehension_orale': row.get(mapping['compr_oral'], '') if mapping.get('compr_oral') else '',
                        'Comprehension_ecrit': row.get(mapping['compr_ecrit'], '') if mapping.get('compr_ecrit') else '',
                        'Production_orale': row.get(mapping['prod_oral'], '') if mapping.get('prod_oral') else '',
                        'Production_ecrit': row.get(mapping['prod_ecrit'], '') if mapping.get('prod_ecrit') else '',
                        'Observation': row.get(mapping['observation'], '') if mapping.get('observation') else '',
                        'Niveau_global': row.get(mapping['niveau'], '') if mapping.get('niveau') else ''
                    }
                    # Convertir les types non sérialisables
                    for k, v in ev.items():
                        if pd.isna(v):
                            ev[k] = ''
                        else:
                            # Garder les valeurs simples
                            try:
                                ev[k] = v if isinstance(v, (int, float, str, bool)) else str(v)
                            except Exception:
                                ev[k] = str(v)
                    evaluations_list.append(ev)

            # Joindre la liste au résultat
            result['evaluations'] = evaluations_list
        except Exception as e:
            # En cas d'erreur sur la lecture des évaluations, fournir une liste vide et logger
            result['evaluations'] = []
            logger.warning(f"Impossible de charger les évaluations pour {numero_apprenant}: {e}")
        
        # Ajouter un message approprié selon le statut
        if presence_status == "no_presence":
            result['message'] = self.cleaning_service.error_messages['no_presence']
            result['message_type'] = 'warning'
        elif presence_status == "has_data":
            result['message'] = f"Données complètes : {total_presences} présences sur {nb_jours_uniques} jours"
            result['message_type'] = 'success'
        else:
            result['message'] = "Aucune donnée disponible"
            result['message_type'] = 'error'

        # === Joindre la grille de progression en fonction du dernier niveau d'évaluation ===
        try:
            grilles_path = self.data_folder / 'grilles_progression.json'
            grille_obj = None
            if grilles_path.exists():
                import json
                with open(grilles_path, 'r', encoding='utf-8') as fh:
                    grilles = json.load(fh)

                # Déterminer le niveau depuis la dernière évaluation si présente
                niveau = None
                if 'evaluations' in result and result['evaluations']:
                    # utiliser le dernier élément (supposé être dans l'ordre du fichier)
                    last_eval = result['evaluations'][-1]
                    niveau = last_eval.get('Niveau_global') or last_eval.get('niveau')

                # Sinon, tenter de lire depuis les informations apprenant (champ pertinent)
                if not niveau and 'apprenant' in result and result['apprenant']:
                    # Exemple: si le champ 'Niveau' existe dans l'inscription
                    niveau = result['apprenant'].get('Niveau', '') or result['apprenant'].get('Niveau_global', '')

                if niveau:
                    niveau_key = str(niveau).strip()
                    # Normaliser: majuscule simple
                    if niveau_key.upper() in grilles:
                        grille_obj = grilles[niveau_key.upper()]
                    elif niveau_key in grilles:
                        grille_obj = grilles[niveau_key]

            if grille_obj:
                result['grille_progression'] = grille_obj
            else:
                result['grille_progression'] = {}
        except Exception as e:
            logger.warning(f"Impossible de charger les grilles de progression: {e}")
            result['grille_progression'] = {}
        
        return result
    
    def _filter_presences_by_date(self, presences_df, dates_filter):
        """
        Filtre les présences selon une période de dates
        
        Args:
            presences_df (DataFrame): DataFrame des présences
            dates_filter (dict): Dictionnaire avec 'debut' et/ou 'fin'
            
        Returns:
            DataFrame: DataFrame filtré
        """
        if dates_filter is None or presences_df.empty:
            return presences_df
            
        filtered_df = presences_df.copy()
        
        # Convertir la colonne de date si elle existe
        if 'Date du Jour' in filtered_df.columns:
            # Essayer différents formats de date
            date_col = filtered_df['Date du Jour'].copy()
            
            # Convertir les dates au format datetime
            converted_dates = []
            for date_val in date_col:
                if pd.isna(date_val) or date_val == '':
                    converted_dates.append(pd.NaT)
                    continue
                    
                try:
                    # Essayer format DD/MM/YYYY
                    if isinstance(date_val, str) and '/' in date_val:
                        parts = date_val.split('/')
                        if len(parts) == 3:
                            converted_date = pd.to_datetime(f"{parts[2]}-{parts[1]}-{parts[0]}")
                            converted_dates.append(converted_date)
                            continue
                    
                    # Essayer conversion directe
                    converted_dates.append(pd.to_datetime(date_val))
                except:
                    converted_dates.append(pd.NaT)
            
            filtered_df['Date_Converted'] = converted_dates
            
            # Appliquer les filtres
            if 'debut' in dates_filter:
                debut = pd.to_datetime(dates_filter['debut'])
                filtered_df = filtered_df[filtered_df['Date_Converted'] >= debut]
                
            if 'fin' in dates_filter:
                fin = pd.to_datetime(dates_filter['fin'])
                filtered_df = filtered_df[filtered_df['Date_Converted'] <= fin]
                
            # Supprimer la colonne temporaire
            filtered_df = filtered_df.drop('Date_Converted', axis=1)
        
        return filtered_df
    
    def _convertir_duree_en_minutes(self, duree_str):
        """
        Convertit une durée en format texte vers des minutes
        
        Args:
            duree_str (str): Durée au format "HH:MM", "Xh", "XhYY", "X h YY", etc.
            
        Returns:
            int: Nombre de minutes
        """
        if not duree_str or str(duree_str).strip() == '':
            return 0
            
        duree = str(duree_str).lower().strip()
        
        try:
            # Format "HH:MM" ou "H:MM"
            if ':' in duree:
                parts = duree.split(':')
                if len(parts) == 2:
                    heures = int(parts[0])
                    minutes = int(parts[1])
                    return (heures * 60) + minutes
            
            # Format "Xh" ou "XhYY" ou "X h YY"
            match_heure = re.match(r'(\d+)\s*h(?:\s*(\d+))?', duree)
            if match_heure:
                heures = int(match_heure.group(1))
                minutes = int(match_heure.group(2)) if match_heure.group(2) else 0
                return (heures * 60) + minutes
            
            # Format "XX min" ou "XXmin"
            match_min = re.match(r'(\d+)\s*min', duree)
            if match_min:
                return int(match_min.group(1))
            
            # Format décimal "X.Y" (heures décimales)
            try:
                nombre = float(duree)
                return int(nombre * 60)
            except ValueError:
                pass
                
        except (ValueError, AttributeError):
            pass
            
        return 0
    
    def _convertir_minutes_en_heures(self, total_minutes):
        """
        Convertit des minutes en format HHhMM
        
        Args:
            total_minutes (int): Nombre total de minutes
            
        Returns:
            str: Durée au format "XXhYY"
        """
        if total_minutes == 0:
            return '0h00'
            
        heures = total_minutes // 60
        minutes = total_minutes % 60
        
        return f"{heures}h{minutes:02d}"
    
    def get_data_quality_info(self):
        """Retourne les informations sur la qualité des données"""
        return {
            'score': self.data_quality_score,
            'last_cleaning': self.last_cleaning_alerts,
            'total_inscriptions': len(self.inscription_df) if not self.inscription_df.empty else 0,
            'total_presences': len(self.presence_df) if not self.presence_df.empty else 0,
            'apprenants_with_presence': len(set(self.presence_df['Numéro Apprenant'])) if not self.presence_df.empty else 0,
            'summary': self.cleaning_service.get_cleaning_summary(
                self.last_cleaning_alerts['inscription'],
                self.last_cleaning_alerts['presence']
            )
        }
    
    def force_data_reload(self):
        """Force le rechargement et nettoyage des données"""
        print("Rechargement force des donnees...")
        # Forcer le rechargement des données via le DataLoader
        self.data_loader.reload_all()
        # Recharger les données dans le service
        self.load_data()
        return self.get_data_quality_info()
    
    def get_apprenant_presence_chart_data(self, numero_apprenant, dates_filter=None):
        """Retourner les données formatées pour Chart.js avec validation et filtrage optionnel par dates"""
        if self.presence_df.empty:
            return {
                'labels': [], 
                'data': [], 
                'total': 0, 
                'jours_uniques': 0,
                'message': self.cleaning_service.error_messages['no_presence'],
                'status': 'no_data'
            }
        
        presence_apprenant = self.presence_df[self.presence_df['Numéro Apprenant'] == numero_apprenant]
        
        if presence_apprenant.empty:
            return {
                'labels': [], 
                'data': [], 
                'total': 0, 
                'jours_uniques': 0,
                'message': self.cleaning_service.error_messages['no_presence'],
                'status': 'no_presence'
            }
        
        # Appliquer le filtrage par dates si demandé
        if dates_filter:
            presence_apprenant = self._filter_presences_by_date(presence_apprenant, dates_filter)
            
            if presence_apprenant.empty:
                return {
                    'labels': [], 
                    'data': [], 
                    'total': 0, 
                    'jours_uniques': 0,
                    'message': 'Aucune présence trouvée pour la période sélectionnée',
                    'status': 'no_data_in_period'
                }
        
        # Grouper par date
        try:
            # Les dates sont déjà nettoyées par le service de nettoyage
            presence_apprenant['Date du Jour'] = pd.to_datetime(presence_apprenant['Date du Jour'], format='%d/%m/%Y', errors='coerce')
            # Supprimer les dates NaT (Not a Time)
            presence_apprenant = presence_apprenant.dropna(subset=['Date du Jour'])
            
            if presence_apprenant.empty:
                return {
                    'labels': [], 
                    'data': [], 
                    'total': 0, 
                    'jours_uniques': 0,
                    'message': 'Données de présence avec dates invalides',
                    'status': 'invalid_dates'
                }
            
            presence_by_date = presence_apprenant.groupby(presence_apprenant['Date du Jour'].dt.date).size()
            
            # Préparer les données pour Chart.js
            labels = [date.strftime('%d/%m/%Y') for date in presence_by_date.index]
            data = [int(count) for count in presence_by_date.values]  # Convertir en int pour éviter les NaN
            
            return {
                'labels': labels,
                'data': data,
                'total': int(len(presence_apprenant)),
                'jours_uniques': int(len(labels)),
                'message': f'Données chargées: {len(labels)} jours de présence',
                'status': 'success'
            }
        except Exception as e:
            print(f"Erreur dans get_apprenant_presence_chart_data: {e}")
            return {
                'labels': [], 
                'data': [], 
                'total': 0, 
                'jours_uniques': 0, 
                'error': str(e),
                'message': 'Erreur lors du traitement des données',
                'status': 'error'
            }

    def get_presence_par_jour_semaine(self, numero_apprenant, dates_filter=None):
        """Retourner la répartition de présence par jour de la semaine avec validation et filtrage optionnel par dates"""
        if self.presence_df.empty:
            return {
                'message': self.cleaning_service.error_messages['no_presence'],
                'status': 'no_data'
            }
        
        presence_apprenant = self.presence_df[self.presence_df['Numéro Apprenant'] == numero_apprenant]
        
        if presence_apprenant.empty:
            return {
                'message': self.cleaning_service.error_messages['no_presence'],
                'status': 'no_presence'
            }
        
        # Appliquer le filtrage par dates si demandé
        if dates_filter:
            presence_apprenant = self._filter_presences_by_date(presence_apprenant, dates_filter)
            
            if presence_apprenant.empty:
                return {
                    'message': 'Aucune présence trouvée pour la période sélectionnée',
                    'status': 'no_data_in_period'
                }
        
        try:
            # Les dates sont déjà nettoyées au format DD/MM/YYYY
            presence_apprenant['Date du Jour'] = pd.to_datetime(presence_apprenant['Date du Jour'], format='%d/%m/%Y', errors='coerce')
            presence_apprenant = presence_apprenant.dropna(subset=['Date du Jour'])
            
            if presence_apprenant.empty:
                return {
                    'message': 'Dates de présence invalides',
                    'status': 'invalid_dates'
                }
            
            presence_apprenant['Jour_Semaine'] = presence_apprenant['Date du Jour'].dt.day_name()
            
            # Traduction des jours en français
            jours_translation = {
                'Monday': 'Lundi',
                'Tuesday': 'Mardi', 
                'Wednesday': 'Mercredi',
                'Thursday': 'Jeudi',
                'Friday': 'Vendredi',
                'Saturday': 'Samedi',
                'Sunday': 'Dimanche'
            }
            
            presence_apprenant['Jour_Semaine_FR'] = presence_apprenant['Jour_Semaine'].map(jours_translation)
            presence_par_jour = presence_apprenant['Jour_Semaine_FR'].value_counts()
            
            result = presence_par_jour.to_dict()
            result['status'] = 'success'
            result['message'] = f'Répartition calculée sur {len(presence_apprenant)} présences'
            
            return result
        except Exception as e:
            print(f"Erreur dans get_presence_par_jour_semaine: {e}")
            return {
                'message': f'Erreur de traitement: {str(e)}',
                'status': 'error'
            }

    def get_evolution_presence_periode(self, numero_apprenant, dates_filter=None):
        """Retourner l'évolution de la présence sur toute la période de formation avec filtrage optionnel par dates"""
        if self.presence_df.empty:
            return {'labels': [], 'data': [], 'moyenne_mobile': [], 'periodes': []}
        
        presence_apprenant = self.presence_df[self.presence_df['Numéro Apprenant'] == numero_apprenant]
        
        if presence_apprenant.empty:
            return {'labels': [], 'data': [], 'moyenne_mobile': [], 'periodes': []}
        
        # Appliquer le filtrage par dates si demandé
        if dates_filter:
            presence_apprenant = self._filter_presences_by_date(presence_apprenant, dates_filter)
            
            if presence_apprenant.empty:
                return {'labels': [], 'data': [], 'moyenne_mobile': [], 'periodes': []}
        
        try:
            # Convertir les dates
            presence_apprenant['Date du Jour'] = pd.to_datetime(presence_apprenant['Date du Jour'], errors='coerce')
            presence_apprenant = presence_apprenant.dropna(subset=['Date du Jour'])
            
            if presence_apprenant.empty:
                return {'labels': [], 'data': [], 'moyenne_mobile': [], 'periodes': []}
            
            # Obtenir la plage complète de dates (du premier au dernier jour)
            date_debut = presence_apprenant['Date du Jour'].min().date()
            date_fin = presence_apprenant['Date du Jour'].max().date()
            
            # Créer une série de toutes les dates dans la période
            toutes_dates = pd.date_range(start=date_debut, end=date_fin, freq='D')
            
            # Compter les présences par date
            presence_by_date = presence_apprenant.groupby(presence_apprenant['Date du Jour'].dt.date).size()
            
            # Créer un DataFrame avec toutes les dates et remplir les manquantes avec 0
            evolution_data = []
            labels = []
            
            for date in toutes_dates:
                date_str = date.strftime('%Y-%m-%d')
                labels.append(date_str)
                count = presence_by_date.get(date.date(), 0)
                evolution_data.append(int(count))
            
            # Calculer une moyenne mobile sur 7 jours pour lisser la courbe
            moyenne_mobile = []
            window_size = 7
            
            for i in range(len(evolution_data)):
                start_idx = max(0, i - window_size // 2)
                end_idx = min(len(evolution_data), i + window_size // 2 + 1)
                moyenne = sum(evolution_data[start_idx:end_idx]) / (end_idx - start_idx)
                moyenne_mobile.append(round(moyenne, 2))
            
            # Identifier les périodes de forte/faible présence
            periodes = self._identifier_periodes_presence(evolution_data, labels)
            
            return {
                'labels': labels,
                'data': evolution_data,
                'moyenne_mobile': moyenne_mobile,
                'periodes': periodes,
                'date_debut': date_debut.strftime('%d/%m/%Y'),
                'date_fin': date_fin.strftime('%d/%m/%Y'),
                'duree_formation': (date_fin - date_debut).days + 1
            }
            
        except Exception as e:
            print(f"Erreur dans get_evolution_presence_periode: {e}")
            return {'labels': [], 'data': [], 'moyenne_mobile': [], 'periodes': [], 'error': str(e)}
    
    def _identifier_periodes_presence(self, data, labels):
        """Identifier les périodes de forte et faible présence"""
        if not data:
            return []
        
        moyenne_generale = sum(data) / len(data)
        periodes = []
        
        # Seuils pour définir les périodes
        seuil_forte = moyenne_generale * 1.5
        seuil_faible = moyenne_generale * 0.5
        
        periode_actuelle = None
        debut_periode = None
        
        for i, (valeur, label) in enumerate(zip(data, labels)):
            if valeur >= seuil_forte and periode_actuelle != 'forte':
                # Début d'une période de forte présence
                if periode_actuelle and debut_periode:
                    periodes.append({
                        'type': periode_actuelle,
                        'debut': debut_periode,
                        'fin': labels[i-1] if i > 0 else label
                    })
                periode_actuelle = 'forte'
                debut_periode = label
                
            elif valeur <= seuil_faible and periode_actuelle != 'faible':
                # Début d'une période de faible présence
                if periode_actuelle and debut_periode:
                    periodes.append({
                        'type': periode_actuelle,
                        'debut': debut_periode,
                        'fin': labels[i-1] if i > 0 else label
                    })
                periode_actuelle = 'faible'
                debut_periode = label
        
        # Fermer la dernière période
        if periode_actuelle and debut_periode:
            periodes.append({
                'type': periode_actuelle,
                'debut': debut_periode,
                'fin': labels[-1]
            })
        
        return periodes

    def add_presence(self, data):
        """Ajouter une nouvelle présence au fichier Excel"""
        try:
            file_path = self.data_folder / 'presence.xlsx'
            
            if file_path.exists():
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                # Créer les en-têtes selon la structure Excel de présence
                headers = [
                    "Date du Jour", "Numéro Apprenant", "Apprenant", 
                    "Activités Apprenants Début", "Activités Apprenants Fin", 
                    "Durée Activité Apprenants", "Activités", 
                    "Activité Encadrant Début", "Activité Encadrant Fin", 
                    "Activité Encadrant Durée", "Numéro Encadrant", 
                    "Encadrant", "Tierce Personne", "Détail"
                ]
                ws.append(headers)
            
            # Préparer les données dans l'ordre des colonnes Excel
            row_data = [
                data.get('date_jour', ''),
                data.get('numero_apprenant', ''),
                data.get('apprenant', ''),
                data.get('activite_debut', ''),
                data.get('activite_fin', ''),
                data.get('duree_activite', ''),
                data.get('activites', ''),
                '',  # Activité Encadrant Début
                '',  # Activité Encadrant Fin
                '',  # Activité Encadrant Durée
                '',  # Numéro Encadrant
                data.get('encadrant', ''),
                '',  # Tierce Personne
                data.get('detail', '')
            ]
            
            ws.append(row_data)
            wb.save(file_path)
            self.load_data()  # Reload data after update
            return True, "Présence ajoutée avec succès !"
        except Exception as e:
            return False, f"Erreur lors de l'ajout : {str(e)}"

    def get_ateliers_par_semaine(self, semaine):
        """Récupérer les ateliers d'une semaine donnée"""
        try:
            # Pour l'instant, on simule avec un fichier JSON ou on crée une structure
            # Plus tard, on pourra utiliser une vraie base de données
            planning_file = self.data_folder / f'planning_{semaine}.json'
            
            if planning_file.exists():
                import json
                with open(planning_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return []
        except Exception as e:
            print(f"Erreur lors du chargement des ateliers: {e}")
            return []

    def creer_atelier(self, data):
        """Créer un nouvel atelier"""
        try:
            import json
            from datetime import datetime
            
            # Générer un ID unique
            atelier_id = int(datetime.now().timestamp() * 1000)
            data['id'] = atelier_id
            data['date_creation'] = datetime.now().isoformat()
            
            # Charger les ateliers existants
            semaine = data['semaine']
            planning_file = self.data_folder / f'planning_{semaine}.json'
            
            ateliers = []
            if planning_file.exists():
                with open(planning_file, 'r', encoding='utf-8') as f:
                    ateliers = json.load(f)
            
            # Ajouter le nouvel atelier
            ateliers.append(data)
            
            # Sauvegarder
            with open(planning_file, 'w', encoding='utf-8') as f:
                json.dump(ateliers, f, ensure_ascii=False, indent=2)
            
            return atelier_id
        except Exception as e:
            raise Exception(f"Erreur lors de la création de l'atelier: {str(e)}")

    def modifier_atelier(self, data):
        """Modifier un atelier existant"""
        try:
            import json
            from datetime import datetime
            
            data['date_modification'] = datetime.now().isoformat()
            
            # Charger les ateliers existants
            semaine = data['semaine']
            planning_file = self.data_folder / f'planning_{semaine}.json'
            
            if not planning_file.exists():
                raise Exception("Planning non trouvé")
            
            with open(planning_file, 'r', encoding='utf-8') as f:
                ateliers = json.load(f)
            
            # Trouver et modifier l'atelier
            atelier_trouve = False
            for i, atelier in enumerate(ateliers):
                if atelier.get('id') == data['id']:
                    ateliers[i] = data
                    atelier_trouve = True
                    break
            
            if not atelier_trouve:
                raise Exception("Atelier non trouvé")
            
            # Sauvegarder
            with open(planning_file, 'w', encoding='utf-8') as f:
                json.dump(ateliers, f, ensure_ascii=False, indent=2)
            
            return data['id']
        except Exception as e:
            raise Exception(f"Erreur lors de la modification de l'atelier: {str(e)}")

    def supprimer_atelier(self, atelier_id):
        """Supprimer un atelier"""
        try:
            import json
            
            # Chercher dans tous les plannings (on ne connaît pas la semaine)
            for planning_file in self.data_folder.glob('planning_*.json'):
                with open(planning_file, 'r', encoding='utf-8') as f:
                    ateliers = json.load(f)
                
                # Filtrer l'atelier à supprimer
                ateliers_filtres = [a for a in ateliers if a.get('id') != atelier_id]
                
                if len(ateliers_filtres) != len(ateliers):
                    # L'atelier a été trouvé et supprimé
                    with open(planning_file, 'w', encoding='utf-8') as f:
                        json.dump(ateliers_filtres, f, ensure_ascii=False, indent=2)
                    return True
            
            raise Exception("Atelier non trouvé")
        except Exception as e:
            raise Exception(f"Erreur lors de la suppression de l'atelier: {str(e)}")

    def get_horaires_planning_defaut(self):
        """Récupérer les horaires par défaut du planning"""
        try:
            import json
            config_file = self.data_folder / 'config_planning.json'
            
            if config_file.exists():
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get('horaires_defaut', [])
            
            # Horaires par défaut si pas de configuration
            return [
                '08:00', '08:30', '09:00', '09:30', '10:00', '10:30',
                '11:00', '11:30', '12:00', '12:30', '13:00', '13:30',
                '14:00', '14:30', '15:00', '15:30', '16:00', '16:30', '17:00'
            ]
        except Exception as e:
            print(f"Erreur lors du chargement des horaires: {e}")
            return []

    def modifier_horaires_planning_defaut(self, horaires):
        """Modifier les horaires par défaut du planning"""
        try:
            import json
            config_file = self.data_folder / 'config_planning.json'
            
            config = {}
            if config_file.exists():
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            config['horaires_defaut'] = horaires
            
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            return True
        except Exception as e:
            raise Exception(f"Erreur lors de la modification des horaires: {str(e)}")

    def exporter_planning_excel(self, semaine):
        """Exporter le planning d'une semaine en Excel"""
        try:
            import json
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Charger les ateliers de la semaine
            ateliers = self.get_ateliers_par_semaine(semaine)
            
            # Créer le workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Planning {semaine}"
            
            # En-têtes
            jours = ['Heure', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi']
            for col, jour in enumerate(jours, 1):
                cell = ws.cell(row=1, column=col, value=jour)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Horaires
            horaires = self.get_horaires_planning_defaut()
            for row, heure in enumerate(horaires, 2):
                ws.cell(row=row, column=1, value=heure)
            
            # Remplir les ateliers
            for atelier in ateliers:
                jour_col = jours.index(atelier['jour'].capitalize()) if atelier['jour'].capitalize() in jours else None
                if jour_col:
                    heure_row = horaires.index(atelier['heure_debut']) + 2 if atelier['heure_debut'] in horaires else None
                    if heure_row:
                        cell_value = f"{atelier['activite']}\n{atelier['encadrant']}\n({len(atelier.get('apprenants', []))} participants)"
                        ws.cell(row=heure_row, column=jour_col, value=cell_value)
            
            # Sauvegarder
            filename = f"planning_{semaine}.xlsx"
            filepath = self.data_folder / filename
            wb.save(filepath)
            
            return filename
        except Exception as e:
            raise Exception(f"Erreur lors de l'export Excel: {str(e)}")

    def dupliquer_planning_semaine(self, semaine_source, semaine_cible):
        """Dupliquer un planning d'une semaine vers une autre"""
        try:
            import json
            
            # Charger le planning source
            ateliers_source = self.get_ateliers_par_semaine(semaine_source)
            
            if not ateliers_source:
                raise Exception("Aucun atelier trouvé dans la semaine source")
            
            # Dupliquer les ateliers
            ateliers_cible = []
            for atelier in ateliers_source:
                nouvel_atelier = atelier.copy()
                nouvel_atelier['id'] = int(datetime.now().timestamp() * 1000)
                nouvel_atelier['semaine'] = semaine_cible
                nouvel_atelier['date_creation'] = datetime.now().isoformat()
                ateliers_cible.append(nouvel_atelier)
            
            # Sauvegarder le planning cible
            planning_file = self.data_folder / f'planning_{semaine_cible}.json'
            with open(planning_file, 'w', encoding='utf-8') as f:
                json.dump(ateliers_cible, f, ensure_ascii=False, indent=2)
            
            return len(ateliers_cible)
        except Exception as e:
            raise Exception(f"Erreur lors de la duplication: {str(e)}")

    def sauvegarder_presences_journee(self, date_jour, presences_data, observations_data=None):
        """Sauvegarder les présences d'une journée dans le fichier Excel avec observations et formatage"""
        try:
            import json
            from datetime import datetime, timedelta
            from openpyxl.styles import Alignment, Font, Border, Side
            
            # Charger les ateliers pour récupérer les informations complètes
            date_obj = datetime.strptime(date_jour, '%Y-%m-%d')
            lundi = date_obj - timedelta(days=date_obj.weekday())
            cleSemaine = f"{lundi.year}-{lundi.month}-{lundi.day}"
            
            ateliers = self.get_ateliers_par_semaine(cleSemaine)
            
            # Essayer aussi avec les classes si pas d'ateliers trouvés
            if not ateliers:
                classes = self.get_classes()
                ateliers = classes  # Utiliser les classes comme ateliers
            
            # Charger le fichier de présences Excel
            file_path = self.data_folder / 'presence.xlsx'
            
            if file_path.exists():
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
            else:
                # Créer un nouveau fichier s'il n'existe pas
                wb = openpyxl.Workbook()
                ws = wb.active
                # Ajouter les en-têtes avec formatage
                headers = [
                    'Date du Jour', 'Numéro Apprenant', 'Apprenant', 
                    'Activités Apprenants Début', 'Activités Apprenants Fin',
                    'Durée Activité Apprenants', 'Activités', 'Activité Encadrant Début', 
                    'Activité Encadrant Fin', 'Activité Encadrant Durée', 'Numéro Encadrant',
                    'Encadrant', 'Tierce Personne', 'Détail'
                ]
                ws.append(headers)
                
                # Formater les en-têtes
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Définir les styles de formatage avec marges haut/bas
            alignment_center = Alignment(
                horizontal='center', 
                vertical='center',
                wrap_text=True,
                indent=0
            )
            alignment_left = Alignment(horizontal='left', vertical='center')
            
            nombre_presences = 0
            observations_data = observations_data or {}
            
            # Pour chaque atelier avec des présences
            for atelier_id, apprenants_presences in presences_data.items():
                # Retirer le préfixe "classe_" si présent pour correspondre aux IDs des classes
                classe_id_str = atelier_id.replace('classe_', '') if atelier_id.startswith('classe_') else atelier_id
                
                # Trouver l'atelier correspondant
                atelier = next((a for a in ateliers if str(a['id']) == str(classe_id_str)), None)
                if not atelier:
                    continue
                
                # Récupérer l'observation pour cette classe (utiliser l'ID original avec préfixe)
                observation_classe = observations_data.get(str(atelier_id), "")
                
                # Pour chaque apprenant avec un statut défini
                for numero_apprenant, statut in apprenants_presences.items():
                    if statut == 'present':  # Ne sauvegarder que les présents
                        # Trouver le nom complet de l'apprenant
                        apprenant_info = self.get_apprenant_by_numero(numero_apprenant)
                        nom_complet = apprenant_info.get('nom_complet', f'Apprenant {numero_apprenant}') if apprenant_info else f'Apprenant {numero_apprenant}'
                        
                        # Calculer la durée
                        duree = self.calculer_duree_atelier(atelier['heure_debut'], atelier['heure_fin'])
                        
                        # Récupérer le numéro encadrant
                        numero_encadrant = self.get_numero_encadrant(atelier['encadrant'])
                        
                        # Construire le détail final
                        detail_final = observation_classe if observation_classe else f" - {atelier.get('description', '')}"
                        
                        # Préparer les données de la ligne
                        row_data = [
                            date_obj.strftime('%d/%m/%Y'),  # Date du Jour (format français)
                            numero_apprenant,  # Numéro Apprenant
                            nom_complet,  # Apprenant
                            atelier['heure_debut'],  # Activités Apprenants Début
                            atelier['heure_fin'],  # Activités Apprenants Fin
                            duree,  # Durée Activité Apprenants
                            atelier['activite'],  # Activités
                            atelier['heure_debut'],  # Activité Encadrant Début
                            atelier['heure_fin'],  # Activité Encadrant Fin
                            duree,  # Activité Encadrant Durée
                            numero_encadrant,  # Numéro Encadrant (MODIFIÉ: récupéré depuis config)
                            atelier['encadrant'],  # Encadrant
                            '',  # Tierce Personne
                            detail_final  # Détail avec observation ou par défaut
                        ]
                        
                        # Ajouter la ligne et formater immédiatement
                        row_num = ws.max_row + 1
                        for col_num, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            
                            # 🎨 NOUVEAU : Appliquer l'alignement centre à TOUTES les colonnes
                            cell.alignment = alignment_center
                        
                        # 🎨 NOUVEAU : Définir la hauteur de ligne pour créer des marges visuelles
                        ws.row_dimensions[row_num].height = 20
                        
                        nombre_presences += 1
            
            # Ajuster automatiquement la largeur des colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Limiter à 50 caractères max
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Sauvegarder le fichier
            wb.save(file_path)
            
            # Recharger les données
            self.load_data()
            
            # Retourner le résultat au format attendu par le contrôleur
            return {
                'nombre_presences': nombre_presences,
                'message': f"{nombre_presences} présences sauvegardées avec succès",
                'apprenants_ignores': []  # Pas de gestion des doublons avec openpyxl pour l'instant
            }
            
        except Exception as e:
            raise Exception(f"Erreur lors de la sauvegarde des présences: {str(e)}")

    def calculer_duree_atelier(self, heure_debut, heure_fin):
        """Calculer la durée d'un atelier"""
        try:
            from datetime import datetime, timedelta
            
            # Parser les heures
            debut = datetime.strptime(heure_debut, '%H:%M')
            fin = datetime.strptime(heure_fin, '%H:%M')
            
            # Calculer la différence
            duree = fin - debut
            
            # Convertir en format HH:MM (avec zéro initial pour les heures)
            total_minutes = int(duree.total_seconds() / 60)
            heures = total_minutes // 60
            minutes = total_minutes % 60
            
            return f"{heures:02d}:{minutes:02d}"  # MODIFIÉ: Format HH:MM avec zéro initial
            
        except Exception as e:
            print(f"Erreur lors du calcul de durée: {e}")
            return "01:30"  # MODIFIÉ: Valeur par défaut avec format HH:MM

    def get_apprenant_by_numero(self, numero):
        """Récupérer les informations d'un apprenant par son numéro"""
        try:
            if self.inscription_df.empty:
                return None
            
            # Chercher l'apprenant par numéro
            apprenant = self.inscription_df[self.inscription_df['N°'].astype(str) == str(numero)]
            
            if not apprenant.empty:
                row = apprenant.iloc[0]
                return {
                    'numero': str(row['N°']),
                    'nom_complet': f"{row.get('NOM', '')} {row.get('Prénom', '')}".strip(),  # MODIFIÉ: NOM Prénom
                    'prenom': row.get('Prénom', ''),
                    'nom': row.get('NOM', ''),
                    'email': row.get('Email', ''),
                    'telephone': row.get('Téléphone', '')
                }
            
            return None
            
        except Exception as e:
            print(f"Erreur lors de la recherche de l'apprenant: {e}")
            return None

    def get_classes(self):
        """Récupérer toutes les classes définies"""
        try:
            import json
            classes_file = self.data_folder / 'classes.json'
            
            if classes_file.exists():
                with open(classes_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                return []
        except Exception as e:
            print(f"Erreur lors du chargement des classes: {e}")
            return []

    def creer_classe(self, data):
        """Créer une nouvelle classe avec verrouillage sécurisé"""
        try:
            from datetime import datetime
            
            # Générer un ID unique
            classe_id = int(datetime.now().timestamp() * 1000)
            data['id'] = classe_id
            data['date_creation'] = datetime.now().isoformat()
            
            # S'assurer que le champ apprenants existe, même vide
            if 'apprenants' not in data:
                data['apprenants'] = []
            
            # Utiliser le service de verrouillage pour une opération atomique
            classes_file = self.data_folder / 'classes.json'
            
            def add_class(classes):
                """Fonction pour ajouter la classe aux données existantes"""
                classes.append(data)
                return classes
            
            # Mise à jour atomique avec verrou
            success = file_lock_service.safe_json_update(classes_file, add_class)
            
            if not success:
                raise Exception("Impossible d'écrire dans le fichier classes.json")
            
            logger.info(f"Classe créée avec succès: {data['nom']} (ID: {classe_id})")
            return classe_id
        except Exception as e:
            raise Exception(f"Erreur lors de la création de la classe: {str(e)}")

    def modifier_classe(self, data):
        """Modifier une classe existante avec verrouillage sécurisé"""
        try:
            from datetime import datetime
            
            data['date_modification'] = datetime.now().isoformat()
            
            # Charger les classes existantes
            classes_file = self.data_folder / 'classes.json'
            
            if not classes_file.exists():
                raise Exception("Fichier des classes non trouvé")
            
            def update_class(classes):
                """Fonction pour modifier la classe dans les données"""
                # Convertir les IDs en string pour une comparaison fiable
                classe_id_recherche = str(data['id'])
                classe_trouvee = False
                
                for i, classe in enumerate(classes):
                    if str(classe.get('id')) == classe_id_recherche:
                        # Conserver l'ID original (numérique) de la classe
                        data['id'] = classe['id']
                        
                        # Préserver les apprenants existants si ils ne sont pas fournis
                        if 'apprenants' not in data or data['apprenants'] is None:
                            data['apprenants'] = classe.get('apprenants', [])
                        
                        # Préserver la date de création si elle existe
                        if 'date_creation' not in data and 'date_creation' in classe:
                            data['date_creation'] = classe['date_creation']
                        
                        classes[i] = data
                        classe_trouvee = True
                        break
                
                if not classe_trouvee:
                    raise Exception(f"Classe non trouvée (ID recherché: {classe_id_recherche})")
                
                return classes
            
            # Mise à jour atomique avec verrou
            success = file_lock_service.safe_json_update(classes_file, update_class)
            
            if not success:
                raise Exception("Impossible d'écrire dans le fichier classes.json")
            
            logger.info(f"Classe modifiée avec succès: {data.get('nom', 'N/A')} (ID: {data['id']})")
            return data['id']
        except Exception as e:
            raise Exception(f"Erreur lors de la modification de la classe: {str(e)}")

    def supprimer_classe(self, classe_id):
        """Supprimer une classe"""
        try:
            import json
            
            classes_file = self.data_folder / 'classes.json'
            
            if not classes_file.exists():
                raise Exception("Fichier des classes non trouvé")
            
            with open(classes_file, 'r', encoding='utf-8') as f:
                classes = json.load(f)
            
            # Filtrer la classe à supprimer
            classes_filtrees = [c for c in classes if c.get('id') != classe_id]
            
            if len(classes_filtrees) == len(classes):
                raise Exception("Classe non trouvée")
            
            # Sauvegarder
            with open(classes_file, 'w', encoding='utf-8') as f:
                json.dump(classes_filtrees, f, ensure_ascii=False, indent=2)
            
            return True
        except Exception as e:
            raise Exception(f"Erreur lors de la suppression de la classe: {str(e)}")

    # Méthode de génération automatique supprimée (nettoyage)

    def get_numero_encadrant(self, nom_encadrant):
        """Récupérer le numéro d'un encadrant par son nom"""
        try:
            config_encadrants = self.get_config_encadrants()
            return config_encadrants.get(nom_encadrant, '')
        except Exception as e:
            print(f"Erreur lors de la récupération du numéro encadrant: {e}")
            return ''

    def get_config_encadrants(self):
        """
        Récupérer la configuration des encadrants depuis config_planning.json
        
        Returns:
            dict: Mapping nom_encadrant -> numero_encadrant
        """
        try:
            import json
            
            config_file = self.data_folder / 'config_planning.json'
            
            if not config_file.exists():
                return {}
            
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            return config.get('encadrants', {})
        except Exception as e:
            print(f"Erreur lors du chargement des encadrants: {e}")
            return {}

    def add_encadrant_mapping(self, nom_encadrant, numero_encadrant):
        """
        Ajouter ou modifier un mapping encadrant dans config_planning.json
        
        Args:
            nom_encadrant (str): Nom de l'encadrant
            numero_encadrant (str): Numéro au format XX-XXX
        
        Returns:
            bool: True si succès, False sinon
        """
        try:
            import json
            
            # Validation du format du numéro
            import re
            if not re.match(r'^\d{2}-\d{3}$', numero_encadrant):
                raise ValueError("Le numéro doit être au format XX-XXX (ex: 25-725)")
            
            # Charger la config actuelle
            config_file = self.data_folder / 'config_planning.json'
            
            if config_file.exists():
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {
                    "horaires_defaut": [],
                    "activites_defaut": [],
                    "encadrants_defaut": [],
                    "encadrants": {}
                }
            
            # Ajouter la section encadrants si elle n'existe pas
            if 'encadrants' not in config:
                config['encadrants'] = {}
            
            # Vérifier que le numéro n'est pas déjà utilisé par un autre encadrant
            for nom_existant, num_existant in config['encadrants'].items():
                if num_existant == numero_encadrant and nom_existant != nom_encadrant:
                    raise ValueError(f"Le numéro {numero_encadrant} est déjà utilisé par {nom_existant}")
            
            # Ajouter/modifier l'encadrant
            config['encadrants'][nom_encadrant] = numero_encadrant
            
            # Sauvegarder
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            return True
            
        except Exception as e:
            raise Exception(f"Erreur lors de l'ajout de l'encadrant: {str(e)}")

    def remove_encadrant_mapping(self, nom_encadrant):
        """
        Supprimer un mapping encadrant de config_planning.json
        
        Args:
            nom_encadrant (str): Nom de l'encadrant à supprimer
        
        Returns:
            bool: True si succès, False sinon
        """
        try:
            import json
            
            config_file = self.data_folder / 'config_planning.json'
            
            if not config_file.exists():
                return False
            
            with open(config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            if 'encadrants' not in config or nom_encadrant not in config['encadrants']:
                return False
            
            del config['encadrants'][nom_encadrant]
            
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            return True
            
        except Exception as e:
            raise Exception(f"Erreur lors de la suppression de l'encadrant: {str(e)}")

    def import_encadrants_from_presence(self):
        """
        Importer automatiquement les mappings encadrants depuis presence.xlsx
        
        Returns:
            dict: Résultat de l'import avec statistiques
        """
        try:
            import json
            
            # Charger presence.xlsx
            presence_file = self.data_folder / 'presence.xlsx'
            if not presence_file.exists():
                return {'success': False, 'error': 'Fichier presence.xlsx non trouvé'}
            
            df_presence = pd.read_excel(presence_file)
            
            # Extraire les mappings uniques encadrant -> numéro
            mappings = df_presence[['Encadrant', 'Numéro Encadrant']].dropna().drop_duplicates()
            
            # Charger la config actuelle
            config_file = self.data_folder / 'config_planning.json'
            if config_file.exists():
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {"encadrants": {}}
            
            if 'encadrants' not in config:
                config['encadrants'] = {}
            
            # Compter les statistiques
            ajoutes = 0
            mises_a_jour = 0
            conflits = 0
            conflits_details = []
            
            for _, row in mappings.iterrows():
                nom = str(row['Encadrant']).strip()
                numero = str(row['Numéro Encadrant']).strip()
                
                if nom and numero and numero != 'nan':
                    # Vérifier les conflits (même numéro, nom différent)
                    conflit_detecte = False
                    for nom_existant, num_existant in config['encadrants'].items():
                        if num_existant == numero and nom_existant != nom:
                            conflits += 1
                            conflits_details.append(f"Numéro {numero}: {nom_existant} vs {nom}")
                            conflit_detecte = True
                            break
                    
                    if not conflit_detecte:
                        if nom in config['encadrants']:
                            if config['encadrants'][nom] != numero:
                                mises_a_jour += 1
                        else:
                            ajoutes += 1
                        
                        config['encadrants'][nom] = numero
            
            # Sauvegarder
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            
            return {
                'success': True,
                'ajoutes': ajoutes,
                'mises_a_jour': mises_a_jour,
                'conflits': conflits,
                'conflits_details': conflits_details
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}

    def ajouter_apprenant_a_classe(self, classe_id, numero_apprenant):
        """Ajouter un apprenant à une classe avec verrouillage sécurisé"""
        try:
            classes_file = self.data_folder / 'classes.json'
            
            if not classes_file.exists():
                return False
            
            def add_student_to_class(classes):
                """Fonction pour ajouter l'apprenant à la classe"""
                classe_trouvee = False
                for classe in classes:
                    if classe['id'] == classe_id:
                        if 'apprenants' not in classe:
                            classe['apprenants'] = []
                        
                        # Vérifier que l'apprenant n'est pas déjà dans la classe
                        if numero_apprenant not in classe['apprenants']:
                            classe['apprenants'].append(numero_apprenant)
                            classe_trouvee = True
                        break
                
                if not classe_trouvee:
                    raise Exception(f"Classe {classe_id} non trouvée")
                
                return classes
            
            # Mise à jour atomique avec verrou
            success = file_lock_service.safe_json_update(classes_file, add_student_to_class)
            
            if success:
                logger.info(f"Apprenant {numero_apprenant} ajouté à la classe {classe_id}")
            
            return success
            
        except Exception as e:
            print(f"Erreur lors de l'ajout de l'apprenant à la classe: {e}")
            return False

    def retirer_apprenant_de_classe(self, classe_id, numero_apprenant):
        """Retirer un apprenant d'une classe"""
        try:
            import json
            classes_file = self.data_folder / 'classes.json'
            
            if not classes_file.exists():
                return False
            
            # Charger les classes
            with open(classes_file, 'r', encoding='utf-8') as f:
                classes = json.load(f)
            
            # Trouver la classe
            classe_trouvee = False
            for classe in classes:
                if classe['id'] == classe_id:
                    if 'apprenants' in classe and numero_apprenant in classe['apprenants']:
                        classe['apprenants'].remove(numero_apprenant)
                        classe_trouvee = True
                    break
            
            if not classe_trouvee:
                return False
            
            # Sauvegarder les modifications
            with open(classes_file, 'w', encoding='utf-8') as f:
                json.dump(classes, f, ensure_ascii=False, indent=2)
            
            return True
            
        except Exception as e:
            print(f"Erreur lors du retrait de l'apprenant de la classe: {e}")
            return False
