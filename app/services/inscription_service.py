# app/services/inscription_service.py
import pandas as pd
import logging
import openpyxl
from openpyxl.styles import Alignment, Font
from datetime import datetime, timedelta
from app.services.base_service import BaseService
from app.utils.validators import validate_inscription_data
from app.utils.helpers import safe_convert_to_int, clean_phone_number
from app.utils.error_handler import handle_errors, DataValidationError

logger = logging.getLogger(__name__)

class InscriptionService(BaseService):
    """
    🎯 Service pour la gestion des inscriptions - VERSION OPTIMISÉE
    
    PROTECTION DU FICHIER SOURCE :
    ===============================
    ✅ TOUTES les opérations de lecture utilisent le cache mémoire (pandas)
    ✅ SEUL l'ajout d'inscription touche physiquement au fichier Excel
    ✅ L'ajout utilise openpyxl pour préserver TOUS les formats (couleurs, largeurs, styles)
    ✅ Les rapports et statistiques N'IMPACTENT JAMAIS le fichier source
    
    PERFORMANCE :
    =============
    ⚡ Cache mémoire pour 2300+ inscriptions = très rapide
    ⚡ Filtrage et tri en mémoire = instantané
    ⚡ Un seul accès fichier lors de l'ajout
    
    FORMATS EXCEL PRÉSERVÉS :
    =========================
    🎨 En-têtes gras avec fond jaune
    🎨 Largeurs de colonnes personnalisées  
    🎨 Tous les styles existants
    """
    
    def __init__(self):
        super().__init__()
        self.required_columns = ['numero_apprenant', 'nom', 'prenom', 'adresse', 'code_postal', 'ville']
    
    @handle_errors(error_type="general")
    def get_total_inscriptions(self):
        """Retourne le nombre total d'inscriptions"""
        count = len(self.inscriptions_df)
        self._log_operation("Comptage inscriptions", f"{count} inscriptions trouvées")
        return count
    
    def generate_next_numero(self):
        """
        Génère le prochain numéro d'inscription disponible
        
        Cette méthode publique utilise la logique interne de génération
        de numéro et peut être appelée depuis les contrôleurs.
        
        Returns:
            str: Le prochain numéro d'inscription au format "AA-NNN"
        """
        try:
            next_numero = self._generate_numero_apprenant()
            self._log_operation("Génération numéro", f"Prochain numéro généré: {next_numero}")
            return next_numero
        except Exception as e:
            logger.error(f"Erreur lors de la génération du numéro: {e}")
            # En cas d'erreur, retourner un numéro basé sur l'année courante
            year = datetime.now().year % 100
            return f"{year:02d}-001"
    
    def _is_numero_apprenant_exists_in_cache(self, numero_apprenant):
        """
        Vérifie si le numéro d'apprenant existe déjà dans le cache mémoire
        PROTECTION : Ne touche jamais au fichier source
        
        Args:
            numero_apprenant (str): Le numéro d'apprenant à vérifier
        Returns:
            bool: True si le numéro existe déjà, False sinon
        """
        if not numero_apprenant:
            return False
        
        try:
            # Utiliser le cache mémoire au lieu de lire le fichier
            if self.inscriptions_df.empty or 'N°' not in self.inscriptions_df.columns:
                return False
            
            # Chercher dans le cache
            existing = self.inscriptions_df[
                self.inscriptions_df['N°'].astype(str).str.strip() == str(numero_apprenant).strip()
            ]
            return len(existing) > 0
            
        except Exception as e:
            logger.error(f"Erreur lors de la vérification du numéro: {e}")
            return False
    
    def _generate_numero_apprenant(self):
        """
        Génère un numéro d'apprenant unique basé sur le cache mémoire
        PROTECTION : Utilise uniquement les données en mémoire
        """
        year = datetime.now().year % 100  # 2 derniers chiffres de l'année
        
        if self.inscriptions_df.empty or 'N°' not in self.inscriptions_df.columns:
            return f"{year:02d}-001"
        
        # Trouver le plus grand numéro existant pour cette année dans le cache
        year_pattern = f"{year:02d}-"
        existing_numbers = self.inscriptions_df[
            self.inscriptions_df['N°'].astype(str).str.startswith(year_pattern, na=False)
        ]['N°'].tolist()
        
        if not existing_numbers:
            return f"{year:02d}-001"
        
        # Extraire les numéros et trouver le maximum
        max_number = 0
        for num in existing_numbers:
            try:
                number_part = int(str(num).split('-')[1])
                max_number = max(max_number, number_part)
            except (ValueError, IndexError):
                continue
        
        next_number = max_number + 1
        return f"{year:02d}-{next_number:03d}"
    
    def get_filtered_inscriptions(self, search='', sexe='', ville='', prioritaire='', pays_naissance='', date_debut='', date_fin='', prescripteur='', structure_actuelle='', mois_anniversaire='', annee_min='', annee_max=''):
        """
        Retourne les inscriptions filtrées selon les critères
        
        Args:
            search (str): Terme de recherche (nom, prénom)
            sexe (str): Filtre par sexe
            ville (str): Filtre par ville
            prioritaire (str): Filtre par statut prioritaire
            pays_naissance (str): Filtre par pays de naissance
            date_debut (str): Date de début au format YYYY-MM-DD
            date_fin (str): Date de fin au format YYYY-MM-DD
            prescripteur (str): Filtre par prescripteur
            structure_actuelle (str): Filtre par structure actuelle
            mois_anniversaire (str): Filtre par mois d'anniversaire (01-12 ou "ce_mois")
            annee_min (str): Année de naissance minimum
            annee_max (str): Année de naissance maximum
            
        Returns:
            list: Liste des inscriptions filtrées
        """
        try:
            df = self.inscriptions_df.copy()
            
            if df.empty:
                return []
            
            # Filtre par recherche textuelle
            if search:
                search_lower = search.lower()
                mask = pd.Series([False] * len(df))
                
                # Recherche dans les colonnes NOM et Prénom (noms Excel corrects)
                if 'NOM' in df.columns:
                    mask |= df['NOM'].astype(str).str.lower().str.contains(search_lower, na=False)
                if 'Prénom' in df.columns:
                    mask |= df['Prénom'].astype(str).str.lower().str.contains(search_lower, na=False)
                if 'N°' in df.columns:
                    mask |= df['N°'].astype(str).str.contains(search, na=False)
                    
                df = df[mask]
            
            # Filtre par sexe
            if sexe:
                if 'Sexe' in df.columns:
                    df = df[df['Sexe'].astype(str).str.lower() == sexe.lower()]
            
            # Filtre par ville
            if ville:
                if 'Ville' in df.columns:
                    df = df[df['Ville'].astype(str).str.lower() == ville.lower()]
            
            # Filtre par prioritaire
            if prioritaire:
                if 'Prioritaire/Veille' in df.columns:
                    if prioritaire.lower() == 'oui':
                        df = df[df['Prioritaire/Veille'].astype(str).str.lower().isin(['oui', 'yes', '1', 'true'])]
                    elif prioritaire.lower() == 'non':
                        df = df[~df['Prioritaire/Veille'].astype(str).str.lower().isin(['oui', 'yes', '1', 'true'])]
            
            # Filtre par pays de naissance
            if pays_naissance:
                if 'Pays de naissance' in df.columns:
                    df = df[df['Pays de naissance'].astype(str).str.lower() == pays_naissance.lower()]

            # Filtre par prescripteur
            if prescripteur:
                if 'Prescripteur' in df.columns:
                    df = df[df['Prescripteur'].astype(str).str.lower() == prescripteur.lower()]

            # Filtre par structure actuelle
            if structure_actuelle:
                if 'Structure actuelle' in df.columns:
                    df = df[df['Structure actuelle'].astype(str).str.lower() == structure_actuelle.lower()]

            # Filtre par mois d'anniversaire
            if mois_anniversaire:
                if 'Date de naissance' in df.columns:
                    try:
                        # Déterminer le mois cible
                        if mois_anniversaire == 'ce_mois':
                            target_month = datetime.now().month
                        else:
                            target_month = int(mois_anniversaire)
                        
                        # Fonction pour extraire le mois de la date de naissance
                        def extract_birth_month(date_str):
                            try:
                                if pd.isna(date_str) or str(date_str).strip() == '':
                                    return None
                                # Gérer différents formats de date
                                date_str = str(date_str).strip()
                                if '/' in date_str:
                                    # Format DD/MM/YYYY
                                    parts = date_str.split('/')
                                    if len(parts) >= 2:
                                        return int(parts[1])  # Mois est la 2ème partie
                                elif '-' in date_str:
                                    # Format YYYY-MM-DD
                                    parts = date_str.split('-')
                                    if len(parts) >= 2:
                                        return int(parts[1])  # Mois est la 2ème partie
                                return None
                            except:
                                return None
                        
                        # Appliquer le filtre
                        birth_months = df['Date de naissance'].apply(extract_birth_month)
                        df = df[birth_months == target_month]
                        
                    except Exception as e:
                        logger.warning(f"Erreur lors du filtrage par mois d'anniversaire: {e}")
                        # En cas d'erreur, ne pas appliquer le filtre

            # Filtre par année de naissance
            if annee_min or annee_max:
                if 'Date de naissance' in df.columns:
                    try:
                        # Fonction pour extraire l'année de la date de naissance
                        def extract_birth_year(date_str):
                            try:
                                if pd.isna(date_str) or str(date_str).strip() == '':
                                    return None
                                date_str = str(date_str).strip()
                                if '/' in date_str:
                                    # Format DD/MM/YYYY
                                    parts = date_str.split('/')
                                    if len(parts) >= 3:
                                        return int(parts[2])  # Année est la 3ème partie
                                elif '-' in date_str:
                                    # Format YYYY-MM-DD
                                    parts = date_str.split('-')
                                    if len(parts) >= 1:
                                        return int(parts[0])  # Année est la 1ère partie
                                return None
                            except:
                                return None
                        
                        # Extraire les années de naissance
                        birth_years = df['Date de naissance'].apply(extract_birth_year)
                        
                        # Appliquer les filtres min et max
                        mask = pd.Series([True] * len(df))
                        
                        if annee_min:
                            min_year = int(annee_min)
                            mask &= (birth_years >= min_year)
                        
                        if annee_max:
                            max_year = int(annee_max)
                            mask &= (birth_years <= max_year)
                        
                        # Filtrer les lignes avec des années valides uniquement
                        mask &= birth_years.notna()
                        
                        df = df[mask]
                        
                    except Exception as e:
                        logger.warning(f"Erreur lors du filtrage par année de naissance: {e}")
                        # En cas d'erreur, ne pas appliquer le filtre
            
            # Filtre par plage de dates d'inscription
            if (date_debut or date_fin) and 'Date inscription' in df.columns:
                try:
                    # Convertir la colonne de dates
                    df['Date inscription'] = pd.to_datetime(df['Date inscription'], errors='coerce')
                    
                    if date_debut:
                        date_debut_dt = pd.to_datetime(date_debut)
                        df = df[df['Date inscription'] >= date_debut_dt]
                        
                    if date_fin:
                        date_fin_dt = pd.to_datetime(date_fin)
                        # Ajouter 1 jour pour inclure la date de fin complète
                        date_fin_dt = date_fin_dt + pd.Timedelta(days=1)
                        df = df[df['Date inscription'] < date_fin_dt]
                        
                except Exception as e:
                    logger.warning(f"Erreur lors du filtrage par dates: {e}")
            
            # Trier par date d'inscription si disponible, sinon par nom
            if 'Date inscription' in df.columns:
                # Convertir les dates en format datetime et gérer les erreurs
                try:
                    df['Date inscription'] = pd.to_datetime(df['Date inscription'], errors='coerce')
                    df = df.sort_values('Date inscription', ascending=False, na_position='last')
                except Exception as e:
                    logger.warning(f"Erreur lors du tri par date d'inscription: {e}")
                    # Tri de secours par nom si le tri par date échoue
                    if 'NOM' in df.columns and 'Prénom' in df.columns:
                        df = df.sort_values(['NOM', 'Prénom'])
            else:
                # Utiliser les noms de colonnes Excel corrects
                if 'NOM' in df.columns and 'Prénom' in df.columns:
                    df = df.sort_values(['NOM', 'Prénom'])
            
            # ✨ CORRECTION : Calculer l'âge dynamiquement avant conversion
            if not df.empty:
                df = self._calculate_age_for_display(df)
            
            result = df.to_dict('records')
            
            # Message de log avec détails des filtres
            filters_used = []
            if search: filters_used.append(f"recherche='{search}'")
            if sexe: filters_used.append(f"sexe='{sexe}'")
            if ville: filters_used.append(f"ville='{ville}'")
            if prioritaire: filters_used.append(f"prioritaire='{prioritaire}'")
            if pays_naissance: filters_used.append(f"pays_naissance='{pays_naissance}'")
            if prescripteur: filters_used.append(f"prescripteur='{prescripteur}'")
            if structure_actuelle: filters_used.append(f"structure_actuelle='{structure_actuelle}'")
            if mois_anniversaire: filters_used.append(f"mois_anniversaire='{mois_anniversaire}'")
            if annee_min: filters_used.append(f"annee_min='{annee_min}'")
            if annee_max: filters_used.append(f"annee_max='{annee_max}'")
            if date_debut: filters_used.append(f"date_debut='{date_debut}'")
            if date_fin: filters_used.append(f"date_fin='{date_fin}'")
            
            filters_text = f" avec filtres: {', '.join(filters_used)}" if filters_used else ""
            self._log_operation("Filtrage inscriptions", f"{len(result)} inscriptions après filtrage{filters_text}")
            return result
            
        except Exception as e:
            logger.error(f"Erreur dans get_filtered_inscriptions: {e}")
            # Retourner une liste vide en cas d'erreur pour éviter les problèmes de template
            return []
    
    def get_inscription_statistics(self):
        """Retourne les statistiques des inscriptions"""
        try:
            df = self.inscriptions_df

            if df.empty:
                return {
                    'total': 0,
                    'mois': 0,
                    'jour': 0,
                    'semaine': 0,
                    'prioritaires': 0,
                    'par_sexe': {},
                    'par_ville': {}
                }

            total = len(df)

            # Statistiques temporelles (si Date inscription disponible)
            mois = semaine = jour = 0
            if 'Date inscription' in df.columns:
                now = datetime.now()
                try:
                    # Travailler sur une copie pour ne pas modifier le DF partagé
                    df_time = df.copy()
                    df_time['date_inscription_dt'] = pd.to_datetime(df_time['Date inscription'], errors='coerce')

                    # Inscriptions du mois
                    debut_mois = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                    mois = len(df_time[df_time['date_inscription_dt'] >= debut_mois])

                    # Inscriptions de la semaine
                    debut_semaine = now - timedelta(days=now.weekday())
                    debut_semaine = debut_semaine.replace(hour=0, minute=0, second=0, microsecond=0)
                    semaine = len(df_time[df_time['date_inscription_dt'] >= debut_semaine])

                    # Inscriptions d'aujourd'hui
                    debut_jour = now.replace(hour=0, minute=0, second=0, microsecond=0)
                    jour = len(df_time[df_time['date_inscription_dt'] >= debut_jour])
                except Exception as e:
                    logger.warning(f"Erreur lors du calcul des statistiques temporelles: {e}")
                    mois = semaine = jour = 0

            # Inscriptions prioritaires
            prioritaires = 0
            if 'Prioritaire/Veille' in df.columns:
                try:
                    prioritaires = len(df[df['Prioritaire/Veille'].astype(str).str.lower().isin(['oui', 'yes', '1', 'true'])])
                except Exception as e:
                    logger.warning(f"Erreur lors du calcul des prioritaires: {e}")
                    prioritaires = 0

            # Répartition par sexe
            par_sexe = {}
            if 'Sexe' in df.columns:
                try:
                    par_sexe = df['Sexe'].value_counts().to_dict()
                except Exception as e:
                    logger.warning(f"Erreur lors du calcul par sexe: {e}")
                    par_sexe = {}

            # Répartition par ville
            par_ville = {}
            if 'Ville' in df.columns:
                try:
                    par_ville = df['Ville'].value_counts().head(10).to_dict()
                except Exception as e:
                    logger.warning(f"Erreur lors du calcul par ville: {e}")
                    par_ville = {}

            stats = {
                'total': total,
                'mois': mois,
                'jour': jour,
                'semaine': semaine,
                'prioritaires': prioritaires,
                'par_sexe': par_sexe,
                'par_ville': par_ville
            }

            self._log_operation("Calcul statistiques", f"Stats calculées pour {total} inscriptions")
            return stats

        except Exception as e:
            logger.error(f"Erreur dans get_inscription_statistics: {e}")
            return {
                'total': 0,
                'mois': 0,
                'jour': 0,
                'semaine': 0,
                'prioritaires': 0,
                'par_sexe': {},
                'par_ville': {}
            }
    
    def get_unique_cities(self):
        """Retourne la liste des villes uniques"""
        try:
            if self.inscriptions_df.empty or 'Ville' not in self.inscriptions_df.columns:
                return []
            
            cities = self.inscriptions_df['Ville'].dropna().unique().tolist()
            cities = [city for city in cities if str(city).strip() and str(city).lower() != 'nan']  # Supprimer les vides et NaN
            cities.sort()
            
            self._log_operation("Récupération villes", f"{len(cities)} villes uniques trouvées")
            return cities
        except Exception as e:
            logger.error(f"Erreur dans get_unique_cities: {e}")
            return []
    
    def get_unique_countries(self):
        """Retourne la liste des pays de naissance uniques"""
        try:
            if self.inscriptions_df.empty or 'Pays de naissance' not in self.inscriptions_df.columns:
                return []
            
            countries = self.inscriptions_df['Pays de naissance'].dropna().unique().tolist()
            countries = [country for country in countries if str(country).strip() and str(country).lower() != 'nan']  # Supprimer les vides et NaN
            countries.sort()
            
            self._log_operation("Récupération pays", f"{len(countries)} pays uniques trouvés")
            return countries
        except Exception as e:
            logger.error(f"Erreur dans get_unique_countries: {e}")
            return []

    def get_unique_prescripteurs(self):
        """Retourne la liste des prescripteurs uniques"""
        try:
            if self.inscriptions_df.empty or 'Prescripteur' not in self.inscriptions_df.columns:
                return []
            
            prescripteurs = self.inscriptions_df['Prescripteur'].dropna().unique().tolist()
            prescripteurs = [prescripteur for prescripteur in prescripteurs if str(prescripteur).strip() and str(prescripteur).lower() != 'nan']
            prescripteurs.sort()
            
            self._log_operation("Récupération prescripteurs", f"{len(prescripteurs)} prescripteurs uniques trouvés")
            return prescripteurs
        except Exception as e:
            logger.error(f"Erreur dans get_unique_prescripteurs: {e}")
            return []

    def get_unique_structures(self):
        """Retourne la liste des structures actuelles uniques"""
        try:
            if self.inscriptions_df.empty or 'Structure actuelle' not in self.inscriptions_df.columns:
                return []
            
            structures = self.inscriptions_df['Structure actuelle'].dropna().unique().tolist()
            structures = [structure for structure in structures if str(structure).strip() and str(structure).lower() != 'nan']
            # Convertir toutes les valeurs en chaînes pour éviter les erreurs de comparaison
            structures = [str(structure) for structure in structures]
            structures.sort()
            
            self._log_operation("Récupération structures", f"{len(structures)} structures uniques trouvées")
            return structures
        except Exception as e:
            logger.error(f"Erreur dans get_unique_structures: {e}")
            return []

    def get_birth_years_range(self):
        """Retourne la plage d'années de naissance (min, max)"""
        try:
            if self.inscriptions_df.empty or 'Date de naissance' not in self.inscriptions_df.columns:
                return {'min_year': 1950, 'max_year': 2010}
            
            # Fonction pour extraire l'année de la date de naissance
            def extract_birth_year(date_str):
                try:
                    if pd.isna(date_str) or str(date_str).strip() == '':
                        return None
                    date_str = str(date_str).strip()
                    if '/' in date_str:
                        # Format DD/MM/YYYY
                        parts = date_str.split('/')
                        if len(parts) >= 3:
                            return int(parts[2])  # Année est la 3ème partie
                    elif '-' in date_str:
                        # Format YYYY-MM-DD
                        parts = date_str.split('-')
                        if len(parts) >= 1:
                            return int(parts[0])  # Année est la 1ère partie
                    return None
                except:
                    return None
            
            # Extraire toutes les années valides
            birth_years = self.inscriptions_df['Date de naissance'].apply(extract_birth_year)
            valid_years = birth_years.dropna().astype(int)
            
            if len(valid_years) > 0:
                min_year = int(valid_years.min())
                max_year = int(valid_years.max())
                self._log_operation("Récupération années naissance", f"Plage: {min_year}-{max_year}")
                return {'min_year': min_year, 'max_year': max_year}
            else:
                # Valeurs par défaut si aucune date valide
                return {'min_year': 1950, 'max_year': 2010}
                
        except Exception as e:
            logger.error(f"Erreur dans get_birth_years_range: {e}")
            return {'min_year': 1950, 'max_year': 2010}
    
    @handle_errors(error_type="general")
    def get_inscription_details(self, numero_apprenant):
        """
        Retourne les détails d'une inscription
        
        Args:
            numero_apprenant (str): Numéro de l'apprenant
            
        Returns:
            dict or None: Détails de l'inscription
        """
        try:
            if self.inscriptions_df.empty or 'N°' not in self.inscriptions_df.columns:
                return None
            
            inscription = self.inscriptions_df[
                self.inscriptions_df['N°'].astype(str) == str(numero_apprenant)
            ]
            
            if inscription.empty:
                self._log_operation("Recherche inscription", f"Aucune inscription trouvée pour {numero_apprenant}")
                return None
            
            # ✨ CORRECTION : Calculer l'âge dynamiquement avant extraction
            inscription = self._calculate_age_for_display(inscription)
            
            details = inscription.iloc[0].to_dict()
            
            # Convertir les valeurs NaT et NaN en chaînes vides pour éviter les erreurs JSON
            for key, value in details.items():
                if pd.isna(value) or value is pd.NaT:
                    details[key] = ""
                elif hasattr(value, 'strftime'):  # Si c'est une date
                    try:
                        details[key] = value.strftime('%d/%m/%Y')
                    except:
                        details[key] = str(value) if value is not None else ""
                else:
                    details[key] = str(value) if value is not None else ""
            
            self._log_operation("Recherche inscription", f"Détails trouvés pour {numero_apprenant}")
            return details
        except Exception as e:
            logger.error(f"Erreur dans get_inscription_details: {e}")
            return None
    
    @handle_errors(error_type="general")
    def export_inscriptions(self):
        """
        🛡️ EXPORT SÉCURISÉ : Exporte les inscriptions vers un fichier Excel
        PROTECTION : Utilise uniquement les données du cache, ne touche jamais au fichier source
        
        Returns:
            str: Chemin du fichier exporté
        """
        if self.inscriptions_df.empty:
            raise DataValidationError("Aucune inscription à exporter")
        
        # Créer le dossier d'export s'il n'existe pas
        export_path = self.data_loader.config.BASE_DIR / 'exports' / f'inscriptions_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        export_path.parent.mkdir(exist_ok=True)
        
        # Exporter UNIQUEMENT depuis le cache mémoire (protection du fichier source)
        self.inscriptions_df.to_excel(export_path, index=False)
        self._log_operation("Export inscriptions", f"Export sécurisé créé: {export_path}")
        
        return str(export_path)

    @handle_errors(error_type="general")
    def add_inscription_with_format_preservation(self, inscription_data):
        """
        🎯 NOUVELLE MÉTHODE : Ajoute une inscription EN PRÉSERVANT tous les formats Excel
        - Utilise openpyxl pour conserver styles, couleurs, largeurs
        - Ajoute UNIQUEMENT une ligne à la fin
        - Ne touche à RIEN d'autre dans le fichier
        
        Args:
            inscription_data (dict): Données de l'inscription
            
        Returns:
            tuple: (success, message)
        """
        try:
            # Validation basique des champs obligatoires
            required_fields = ['numero_apprenant', 'nom', 'prenom', 'adresse', 'code_postal', 'ville']
            missing_fields = [field for field in required_fields if not inscription_data.get(field)]
            
            if missing_fields:
                error_msg = f"Champs obligatoires manquants: {', '.join(missing_fields)}"
                self._log_operation("Ajout inscription", f"ERREUR - {error_msg}")
                return False, error_msg

            # Vérifier si le numéro d'apprenant existe déjà (via cache)
            if self._is_numero_apprenant_exists_in_cache(inscription_data['numero_apprenant']):
                # Générer le prochain numéro disponible pour aider l'utilisateur
                next_numero = self._generate_numero_apprenant()
                error_msg = f"Le numéro d'apprenant {inscription_data['numero_apprenant']} existe déjà. Prochain numéro disponible: {next_numero}"
                self._log_operation("Ajout inscription", f"NUMÉRO EXISTANT - {error_msg}")
                return False, error_msg
            
            # ✨ MAGIE : Utiliser openpyxl pour préserver les formats
            excel_path = self.data_loader.config.INSCRIPTION_FILE
            
            # Charger le fichier Excel avec TOUS ses formats
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Préparer les données dans l'ordre exact des colonnes
            excel_row_data = self._prepare_excel_row_data(inscription_data)
            
            # Ajouter la nouvelle ligne à la fin (nouvelle ligne)
            new_row_number = ws.max_row + 1
            
            # Parcourir chaque colonne et ajouter la valeur avec formatage
            for col_num, column_name in enumerate(excel_row_data.keys(), 1):
                cell = ws.cell(row=new_row_number, column=col_num)
                cell.value = excel_row_data[column_name]
                
                # 🎨 NOUVEAU : Appliquer l'alignement centre avec marges haut/bas + GRAS
                cell.alignment = Alignment(
                    horizontal='center', 
                    vertical='center',
                    wrap_text=True,
                    indent=0
                )
                cell.font = Font(bold=False, size=14)

            # 🎨 NOUVEAU : Définir la hauteur de ligne pour créer des marges visuelles
            ws.row_dimensions[new_row_number].height = 40
            
            # Pas de copie de style pour éviter les erreurs - juste ajouter les données
            # Les formats des en-têtes et colonnes sont automatiquement préservés par openpyxl
            
            # Sauvegarder le fichier (préserve TOUS les formats existants)
            wb.save(excel_path)
            
            # Recharger les données dans le cache
            self.data_loader.reload_inscriptions()
            
            success_msg = f"✅ Inscription ajoutée avec succès pour {inscription_data['prenom']} {inscription_data['nom']} (N° {inscription_data['numero_apprenant']})"
            self._log_operation("Ajout inscription", success_msg)
            return True, success_msg
            
        except Exception as e:
            error_msg = f"❌ Erreur lors de l'ajout de l'inscription: {str(e)}"
            logger.error(error_msg, exc_info=True)
            return False, error_msg
    
    def _prepare_excel_row_data(self, data):
        """
        🎯 NOUVELLE MÉTHODE : Prépare les données pour openpyxl
        Retourne un dictionnaire ordonné selon les colonnes Excel exactes
        """
        
        # Calculer l'âge si date de naissance fournie
        age = ""
        if data.get('date_naissance'):
            try:
                date_naissance = datetime.strptime(data['date_naissance'], '%Y-%m-%d').date()
                today = datetime.now().date()
                age_num = today.year - date_naissance.year - ((today.month, today.day) < (date_naissance.month, date_naissance.day))
                age = f"{age_num} ans"
            except:
                age = ""
        
        # Formater la date de naissance pour l'Excel (DD/MM/YYYY)
        date_naissance_formatted = ""
        if data.get('date_naissance'):
            try:
                date_obj = datetime.strptime(data['date_naissance'], '%Y-%m-%d')
                date_naissance_formatted = date_obj.strftime('%d/%m/%Y')
            except:
                date_naissance_formatted = ""
        
        # Formater les autres dates
        def format_date_for_excel(date_str):
            if not date_str:
                return ""
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                return date_obj.strftime('%d/%m/%Y')
            except:
                return ""
        
        # 📋 ORDRE EXACT des colonnes de votre fichier Excel
        return {
            'N°': data.get('numero_apprenant', ''),
            'NOM': data.get('nom', ''),
            'Prénom': data.get('prenom', ''),
            'Sexe': data.get('sexe', ''),
            'Date de naissance': date_naissance_formatted,
            'Age': age,
            'Pays de naissance': data.get('pays_naissance', ''),
            'Nationalité': data.get('nationalite', ''),
            'Continent': data.get('continent', ''),
            'ISO': data.get('iso', '').upper() if data.get('iso') else '',
            'Arrivée en France': format_date_for_excel(data.get('arrivee_france', '')),
            'Document': data.get('document', ''),
            "Statut à l'entrée": data.get('statut_entree', ''),
            'Statut actuel': data.get('statut_actuel', ''),
            'Adresse': data.get('adresse', ''),
            'Code postal': data.get('code_postal', ''),
            'Ville': data.get('ville', ''),
            'Prioritaire/Veille': data.get('prioritaire_veille', ''),
            'Type de logement': data.get('type_logement', ''),
            'Téléphone': data.get('telephone', ''),
            'Email': data.get('email', ''),
            'Situation Familiale': data.get('situation_familiale', ''),
            'Revenus': data.get('revenus', ''),
            'Langues Parlées': data.get('langues_parlees', ''),
            'Prescripteur': data.get('prescripteur', ''),
            'Structure actuelle': data.get('structure_actuelle', ''),
            'Date inscription': format_date_for_excel(data.get('date_inscription', '')),
            'Première venue': format_date_for_excel(data.get('premiere_venue', '')),
            'Commentaires': data.get('commentaires', '')
        }

    def _calculate_age_for_display(self, df):
        """
        ✨ NOUVELLE MÉTHODE : Calcule l'âge dynamiquement pour l'affichage
        
        Cette méthode prend un DataFrame et calcule l'âge en temps réel
        à partir de la colonne 'Date de naissance', en gérant les différents formats
        
        Args:
            df (pandas.DataFrame): DataFrame avec une colonne 'Date de naissance'
            
        Returns:
            pandas.DataFrame: DataFrame avec la colonne 'Age' mise à jour
        """
        if 'Date de naissance' not in df.columns:
            return df
            
        # Créer une copie pour ne pas modifier l'original
        df_copy = df.copy()
        
        def calculate_age_from_birth_date(birth_date_str):
            """Calcule l'âge à partir d'une date de naissance (plusieurs formats supportés)"""
            if pd.isna(birth_date_str) or birth_date_str == '':
                return None
                
            try:
                # Convertir en string au cas où
                birth_str = str(birth_date_str).strip()
                
                # Format français: DD/MM/YYYY
                if '/' in birth_str and len(birth_str.split('/')) == 3:
                    day, month, year = birth_str.split('/')
                    birth_date = datetime(int(year), int(month), int(day)).date()
                
                # Format ISO: YYYY-MM-DD ou YYYY-MM-DD HH:MM:SS
                elif '-' in birth_str:
                    if ' ' in birth_str:  # Format avec heure
                        birth_str = birth_str.split(' ')[0]  # Prendre seulement la date
                    year, month, day = birth_str.split('-')
                    birth_date = datetime(int(year), int(month), int(day)).date()
                
                else:
                    return None
                    
                # Calculer l'âge
                today = datetime.now().date()
                age = today.year - birth_date.year
                
                # Ajuster si l'anniversaire n'est pas encore passé cette année
                if (today.month, today.day) < (birth_date.month, birth_date.day):
                    age -= 1
                    
                return age
                
            except (ValueError, IndexError, AttributeError):
                return None
        
        # Appliquer le calcul d'âge sur toute la colonne
        df_copy['Age'] = df_copy['Date de naissance'].apply(calculate_age_from_birth_date)
        
        # Convertir les âges en format texte pour l'affichage (seulement les valeurs non-nulles)
        df_copy['Age'] = df_copy['Age'].apply(lambda x: f"{int(x)} ans" if pd.notna(x) and x is not None else None)
        
        return df_copy

    # 🔄 MÉTHODE DE COMPATIBILITÉ : Rediriger l'ancienne méthode vers la nouvelle
    def add_inscription_simple(self, inscription_data):
        """
        ⚠️ REDIRECTION : Cette méthode redirige vers la nouvelle méthode qui préserve les formats
        """
        return self.add_inscription_with_format_preservation(inscription_data)